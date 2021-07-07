# import openpyxl
from openpyxl import load_workbook
import os
for i in range(1,9):
	wb= load_workbook((os.path.join(os.getcwd(), "EXPRESS{}.xlsx".format(i))))

	# print("l.27")

	ws = wb.active
	# print("l.29")

	familyCells=ws['E']
	sFamilyCells=ws['F']
	cell=ws['A1']
	print(cell.row)
	# methods=[meth for meth in dir(cell)]
	# print(methods)

	def renaming(cell):
		famCell=ws['E{}'.format(cell.row)] #accessing to cell
		if cell.row==1:
			return 0

		numberOf0=6-(len(str(famCell.value))+len(str(cell.value)))
		# print("numberOf0",numberOf0)
		# if numberOf0>3:
			# print("famCell",famCell.value)
			# print("cell",cell.value)
			# print('numberOf0',numberOf0)
			# input('pause')
		zeros=""
		for i in range(numberOf0):
			zeros+="0"
		
		cell.value=str(famCell.value) + str(zeros) + str(cell.value)

		cell.value=int(cell.value)
		

		if len(str(cell.value))>6:
				print("famCell",famCell.value)
				print("cell",cell.value)
				print("row",cell.row)
				print('numberOf0',numberOf0)
				input("TROUBLE")
				return 0

		# print("val",cell.value)
		return 1 

	l=[renaming(c) for c in sFamilyCells]
	print("ok on a fait le numero {}".format(i))
	wb.save('ESPRESS{}.xlsx'.format(i))



"""
['__class__', '__delattr__', '__dir__', '__doc__', '__eq__', '__format__', '__ge__',
 '__getattribute__', '__gt__', '__hash__', '__init__', '__init_subclass__', '__le__',
  '__lt__', '__module__', '__ne__', '__new__', '__reduce__', '__reduce_ex__', '__repr__', 
  '__setattr__', '__sizeof__', '__slots__', '__str__', '__subclasshook__', '_bind_value',
   '_comment', '_hyperlink', '_style', '_value', 'alignment', 'base_date', 'border', 
   'check_error', 'check_string', 'col_idx', 'column', 'column_letter', 'comment', 'coordinate', 
   'data_type', 'encoding', 'fill', 'font', 'has_style', 'hyperlink', 'internal_value', 'is_date',
   'number_format', 'offset', 'parent', 'pivotButton', 'protection', 'quotePrefix', 'row', 'style',
    'style_id', 'value']  
"""