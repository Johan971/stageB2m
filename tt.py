# CONSIDERATIONS : Colonne en format fractions dans excel (je parle du fichier .csv)
from openpyxl import load_workbook
import os, pyexcel
import pandas as pd

# sheet = pyexcel.get_sheet(file_name="ARTICLE.CSV", delimiter=";")
# sheet.save_as("ARTICLE.xlsx")
# input("ok chelou ")
wb= load_workbook((os.path.join(os.getcwd(), "ARTICLE.xlsx")))

# print("l.27")

ws = wb.active
# print("l.29")

codeEAN=ws['A']
cell=ws['A1']
print(cell.value)
# input("brek1")
# input("okay")
# methods=[meth for meth in dir(cell)]
# print(methods)

def renaming(cell):
	
	if cell.row==1:
		return 0
	elif type(cell.value)!=type(55):
		rowToDel=ws[cell.row]
		for cells in rowToDel:
			cells.value=""
			return 0
	# print("cellvalue", cell.value,"e")
	cell.value=int(cell.value)
	numberOf0=0
	# print("cell : ",cell.value, len(str(cell.value)))
	if (10<len(str(cell.value))<13):
		# print("new",cell.value)

		numberOf0=13-len(str(cell.value))
		# print("numberOf0",numberOf0)
		# if numberOf0>3:
			# print("famCell",famCell.value)
			# print("cell",cell.value)
			# print('numberOf0',numberOf0)
			# input('pause')

		zeros=""
		for i in range(numberOf0):
			zeros+="0"
		
		cell.value=str(zeros)+ str(cell.value)
		print("newvalue : ",cell.value)
		
	return 1 

l=[renaming(c) for c in codeEAN]
# print("ok on a fait le numero {}".format(i))
wb.save('goodArticle.xlsx')

read_file = pd.read_excel ("goodArticle.xlsx")
read_file.to_csv(os.path.join(os.getcwd(), "goodArticle.csv"), index = None, header=True,sep =";")
# input("brek2")




"""
['__class__', '__delattr__', '__dir__', '__doc__', '__eq__', '__format__', '__ge__',
 '__getattribute__', '__gt__', '__hash__', '__init__', '__init_subclass__', '__le__',
  '__lt__', '__module__', '__ne__', '__new__', '__reduce__', '__reduce_ex__', '__repr__', 
  '__setattr__', '__sizeof__', '__slots__', '__str__', '__subclasshook__', '_bind_value',
   '_comment', '_hyperlink', '_style', '_value', 'alignment', 'base_date', 'border', 
   'check_error', 'check_string', 'col_idx', 'column', 'column_letter', 'comment', 'coordinate', 
   'data_type', 'encoding', 'fill', 'font', 'has_style', 'hyperlink', 'internal_value', 'is_date',
   'number_format', 'offset', 'parent', 'pivotButton', 'protection', 'quotePrefix', 'row', 'style',
    'style_id', 'value']  
"""